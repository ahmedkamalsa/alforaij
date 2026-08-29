"""توليد تقرير Excel احترافي للبحث والتقييم العقاري.

- ورقة النتائج الرئيسية مع كل المؤشرات
- ورقة المقارنات التفصيلية
- ورقة مقارنة التمويل (البنوك)
- ورقة التأمين
- تنسيق احترافي مع عناوين وحدود وألوان
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ألوان موحدة مع تصميم الموقع
NAVY = "0F2A4A"
GOLD = "C9A227"
LIGHT_BG = "F4F6F8"
GREEN = "22C55E"
YELLOW = "F59E0B"
RED = "EF4444"
WHITE = "FFFFFF"
GRAY = "5A6472"

_header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
_header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
_gold_fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
_light_fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
_green_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
_yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
_red_fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
_thin_border = Border(
    left=Side(style="thin", color="D7DCE2"),
    right=Side(style="thin", color="D7DCE2"),
    top=Side(style="thin", color="D7DCE2"),
    bottom=Side(style="thin", color="D7DCE2"),
)
_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
_right = Alignment(horizontal="right", vertical="center", wrap_text=True)


def _auto_width(ws, min_width=12, max_width=40):
    """ ضبط عرض الأعمدة تلقائيًا حسب المحتوى."""
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                # تقدير الطول: العربية أعرض
                cjk = sum(1 for c in val if ord(c) > 0x600)
                ascii_count = len(val) - cjk
                cell_len = cjk * 1.8 + ascii_count
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _write_header_row(ws, row: int, headers: list[str]):
    """كتابة صف عنوان بتنسيق احترافي."""
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = _center
        cell.border = _thin_border


def _write_data_row(ws, row: int, values: list[Any], alt=False):
    """كتابة صف بيانات مع تلوين متناوب."""
    fill = _light_fill if alt else PatternFill()
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.alignment = _right
        cell.border = _thin_border
        if alt:
            cell.fill = fill


def _trust_color(score: int) -> PatternFill:
    """لون شارة الثقة حسب الدرجة."""
    if score >= 75:
        return _green_fill
    elif score >= 50:
        return _yellow_fill
    return _red_fill


def build_excel(report: dict | None) -> bytes:
    """توليد ملف Excel من تقرير البحث. يعيد البايتات."""
    report = report or {}
    request = report.get("request") or {}
    results = report.get("results") or []
    ai = report.get("aiInsights") or {}

    wb = Workbook()

    # ═══════════════════════════════════════════
    # ورقة 1: النتائج الرئيسية
    # ═══════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "النتائج الرئيسية"
    ws1.sheet_properties.tabColor = NAVY

    # معلومات الطلب
    ws1.merge_cells("A1:L1")
    title_cell = ws1["A1"]
    title_cell.value = f"تقرير البحث العقاري — {datetime.now():%Y-%m-%d %H:%M}"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color=NAVY)
    title_cell.alignment = _center

    ws1.merge_cells("A2:L2")
    req_text = request.get("raw_text") or request.get("text") or "—"
    ws1["A2"].value = f"طلب العميل: {req_text}"
    ws1["A2"].font = Font(name="Calibri", size=11, color=GRAY)
    ws1["A2"].alignment = _center

    # جدول النتائج
    headers = [
        "#", "كود الإعلان", "المنطقة", "نوع العقار", "السعر (د.ك)",
        "المساحة (م²)", "سعر المتر", "وسيط المقارنات", "فجوة السعر",
        "درجة التوصية", "ثقة الإعلان", "نطاق الثقة",
    ]
    _write_header_row(ws1, 4, headers)

    for i, item in enumerate(results[:30], start=1):
        row = i + 4
        price = item.get("price") or 0
        space = item.get("space") or 0
        sqm = round(price / space) if space else 0
        median = item.get("marketMedian") or 0
        gap_pct = round(((price / median) - 1) * 100, 1) if median else 0
        gap_label = f"{gap_pct:+.1f}%" if median else "—"
        rec = round(item.get("recommendationScore") or 0)
        ts = item.get("trustScore") or {}
        ts_score = ts.get("score", "—")
        ci = item.get("confidenceInterval") or {}
        ci_text = ci.get("display", "—")

        values = [
            i,
            item.get("code") or "—",
            item.get("area") or "—",
            item.get("propertyType") or item.get("detailClass") or "—",
            price,
            space or "غير محدد",
            sqm or "—",
            median or "—",
            gap_label,
            rec,
            ts_score,
            ci_text,
        ]
        _write_data_row(ws1, row, values, alt=(i % 2 == 0))

        # تلوين درجة التوصية
        rec_cell = ws1.cell(row=row, column=10)
        if rec >= 75:
            rec_cell.fill = _green_fill
            rec_cell.font = Font(bold=True, color=WHITE)
        elif rec >= 50:
            rec_cell.fill = _yellow_fill
            rec_cell.font = Font(bold=True, color=NAVY)

        # تلوين ثقة الإعلان
        ts_cell = ws1.cell(row=row, column=11)
        if isinstance(ts_score, (int, float)):
            ts_cell.fill = _trust_color(int(ts_score))
            ts_cell.font = Font(bold=True, color=WHITE)

    _auto_width(ws1)
    ws1.freeze_panes = "A5"

    # ═══════════════════════════════════════════
    # ورقة 2: المقارنات التفصيلية
    # ═══════════════════════════════════════════
    ws2 = wb.create_sheet("المقارنات")
    ws2.sheet_properties.tabColor = GOLD

    comp_headers = [
        "#", "الإعلان الأصلي", "كود المقارن", "المنطقة", "السعر (د.ك)",
        "المساحة (م²)", "سعر المتر", "التاريخ", "المصدر", "الرابط",
    ]
    _write_header_row(ws2, 1, comp_headers)

    comp_row = 2
    for item in results[:20]:
        comps = item.get("comparables") or []
        for j, comp in enumerate(comps[:8]):
            comp_price = comp.get("price") or 0
            comp_space = comp.get("space") or 0
            comp_sqm = round(comp_price / comp_space) if comp_space else 0
            values = [
                comp_row - 1,
                item.get("code") or "—",
                comp.get("code") or "—",
                comp.get("area") or "—",
                comp_price,
                comp_space or "—",
                comp_sqm or "—",
                comp.get("date") or "—",
                comp.get("source") or "—",
                comp.get("url") or comp.get("originalUrl") or "—",
            ]
            _write_data_row(ws2, comp_row, values, alt=(comp_row % 2 == 0))
            comp_row += 1

    _auto_width(ws2)
    ws2.freeze_panes = "A2"

    # ═══════════════════════════════════════════
    # ورقة 3: مقارنة التمويل (البنوك)
    # ═══════════════════════════════════════════
    ws3 = wb.create_sheet("مقارنة البنوك")
    ws3.sheet_properties.tabColor = "1A7F37"

    top = results[0] if results else {}
    price = top.get("price") or 0
    if price > 0:
        try:
            from backend.services.mortgage_calculator import compare_banks
            result = compare_banks(price, 30, 20)
            banks = result.get("banks", [])
            best_code = result.get("best_bank")

            ws3.merge_cells("A1:G1")
            ws3["A1"].value = f"مقارنة التمويل العقاري — عقار بقيمة {price:,.0f} د.ك (30% دفعة مقدمة، 20 سنة)"
            ws3["A1"].font = Font(name="Calibri", size=12, bold=True, color=NAVY)
            ws3["A1"].alignment = _center

            bank_headers = [
                "البنك", "الفائدة (%)", "القسط الشهري (د.ك)", "إجمالي الفائدة (د.ك)",
                "الإجمالي المدفوع (د.ك)", "المدة", "ملاحظات",
            ]
            _write_header_row(ws3, 3, bank_headers)

            for j, bank in enumerate(banks):
                row = j + 4
                is_best = bank.get("code") == best_code
                values = [
                    f"{'🏆 ' if is_best else ''}{bank.get('name', '')}",
                    bank.get("rate", 0),
                    bank.get("monthly_payment", 0),
                    bank.get("total_interest", 0),
                    bank.get("total_paid", 0),
                    f"{bank.get('years', 0)} سنة",
                    "الأفضل" if is_best else "",
                ]
                _write_data_row(ws3, row, values, alt=(j % 2 == 0))
                if is_best:
                    for col in range(1, 8):
                        ws3.cell(row=row, column=col).fill = _green_fill
                        ws3.cell(row=row, column=col).font = Font(bold=True, color=WHITE)

            # التوصية
            rec_data = result.get("recommendation", {})
            if rec_data.get("summary"):
                rec_row = len(banks) + 5
                ws3.merge_cells(f"A{rec_row}:G{rec_row}")
                ws3[f"A{rec_row}"].value = f"التوصية: {rec_data['summary']}"
                ws3[f"A{rec_row}"].font = Font(name="Calibri", size=11, bold=True, color=NAVY)

            _auto_width(ws3)
            ws3.freeze_panes = "A4"
        except Exception:
            ws3["A1"].value = "بيانات التمويل غير متاحة"
    else:
        ws3["A1"].value = "لا يوجد سعر كافٍ لحساب التمويل"

    # ═══════════════════════════════════════════
    # ورقة 4: التأمين
    # ═══════════════════════════════════════════
    ws4 = wb.create_sheet("التأمين")
    ws4.sheet_properties.tabColor = "7C3AED"

    if price > 0:
        try:
            from backend.services.insurance_calculator import calculate_insurance
            ins_result = calculate_insurance(price, contents_value=50000, building_age=5, years=1)
            if ins_result:
                ws4.merge_cells("A1:D1")
                ws4["A1"].value = f"تقدير التأمين العقاري — عقار بقيمة {price:,.0f} د.ك"
                ws4["A1"].font = Font(name="Calibri", size=12, bold=True, color=NAVY)
                ws4["A1"].alignment = _center

                ins_headers = ["نوع التأمين", "القسط السنوي (د.ك)", "القسط الشهري (د.ك)", "الوصف"]
                _write_header_row(ws4, 3, ins_headers)

                for j, ins in enumerate(ins_result.get("types", [])):
                    row = j + 4
                    values = [
                        ins.get("name") or "—",
                        ins.get("annual", 0),
                        ins.get("monthly", 0),
                        ins.get("description") or "—",
                    ]
                    _write_data_row(ws4, row, values, alt=(j % 2 == 0))

                # الخصومات
                discounts = ins_result.get("discounts") or []
                if discounts:
                    disc_row = len(ins_result.get("types", [])) + 5
                    ws4.merge_cells(f"A{disc_row}:D{disc_row}")
                    ws4[f"A{disc_row}"].value = "الخصومات المتاحة:"
                    ws4[f"A{disc_row}"].font = Font(bold=True, color=NAVY)
                    for j, d in enumerate(discounts):
                        r = disc_row + 1 + j
                        ws4.cell(row=r, column=1, value=f"• {d.get('name', '')}")
                        ws4.cell(row=r, column=2, value=f"خصم {d.get('percent', 0)}%")

                _auto_width(ws4)
                ws4.freeze_panes = "A4"
        except Exception:
            ws4["A1"].value = "بيانات التأمين غير متاحة"
    else:
        ws4["A1"].value = "لا يوجد سعر كافٍ لحساب التأمين"

    # ═══════════════════════════════════════════
    # ورقة 5: عوامل التقييم
    # ═══════════════════════════════════════════
    ws5 = wb.create_sheet("عوامل التقييم")
    ws5.sheet_properties.tabColor = "DC2626"

    ef_headers = ["الإعلان", "#", "العامل", "النوع", "التفاصيل"]
    _write_header_row(ws5, 1, ef_headers)

    ef_row = 2
    for item in results[:20]:
        factors = item.get("explanationFactors") or []
        for j, f in enumerate(factors):
            ftype = {"positive": "إيجابي", "negative": "سلبي", "neutral": "محايد", "info": "معلومات"}.get(
                f.get("type", ""), f.get("type", "")
            )
            values = [
                item.get("code") or "—",
                j + 1,
                f"{f.get('icon', '')} {f.get('label', '') or f.get('text', '')}",
                ftype,
                f.get("detail") or f.get("value") or "—",
            ]
            _write_data_row(ws5, ef_row, values, alt=(ef_row % 2 == 0))
            ef_row += 1

    _auto_width(ws5)
    ws5.freeze_panes = "A2"

    # ═══════════════════════════════════════════
    # ورقة 6: ملخص التحليل
    # ═══════════════════════════════════════════
    ws6 = wb.create_sheet("ملخص التحليل")
    ws6.sheet_properties.tabColor = "0EA5E9"

    summary = report.get("summary") or "—"
    ws6.merge_cells("A1:B1")
    ws6["A1"].value = "الخلاصة التنفيذية"
    ws6["A1"].font = Font(name="Calibri", size=14, bold=True, color=NAVY)

    ws6.merge_cells("A2:B2")
    ws6["A2"].value = summary
    ws6["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws6.row_dimensions[2].height = 100

    # اقتراحات AI
    suggestions = ai.get("suggestions") or ""
    if suggestions:
        ws6["A4"].value = "توصيات للمستشار:"
        ws6["A4"].font = Font(bold=True, color=NAVY)
        ws6.merge_cells("A5:B5")
        ws6["A5"].value = suggestions
        ws6["A5"].alignment = Alignment(wrap_text=True, vertical="top")
        ws6.row_dimensions[5].height = 80

    ws6.column_dimensions["A"].width = 30
    ws6.column_dimensions["B"].width = 80

    # حفظ في الذاكرة
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
